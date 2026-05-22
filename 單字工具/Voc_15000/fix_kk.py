import sys
import io
import re
import openpyxl
import json
import csv
import shutil
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Backup first
ts = datetime.now().strftime('%Y%m%d_%H%M%S')
shutil.copy('data/word_list.xlsx', f'backup/word_list_backup_{ts}.xlsx')
print(f'Backup: backup/word_list_backup_{ts}.xlsx')

wb = openpyxl.load_workbook('data/word_list.xlsx')
ws = wb.active
headers = [cell.value for cell in ws[1]]
kk_col = headers.index('kk_phonetics') + 1
word_col = headers.index('word') + 1

fixes = []

# Manual overrides — word -> correct kk
manual = {
    'corporation': 'ˌkɔrpɚ\'eʃən',
    'coalition':   'ˌkoəlɪ\'ʃən',
    'creation':    'kri\'eʃən',
    'jurisdiction':'ˌdʒʊrɪs\'dɪkʃən',
}

def first_vowel_is_full(kk):
    """Return True if the first vowel character is a full (non-reduced) vowel."""
    reduced = set('əɪɚɝ')
    full = set('æɑɔʌɛeioua')
    for c in kk:
        if c in reduced:
            return False
        if c in full:
            return True
    return False

def needs_secondary_stress(word, kk):
    """
    Returns True if this -ation/-tion/-sion/-ution word should get a leading ˌ.
    Conditions:
    - Does NOT already start with ˌ or '
    - Has a ' somewhere (has primary stress marked)
    - first_vowel_is_full() is True
    """
    if not kk:
        return False
    if kk[0] in ('ˌ', "'"):
        return False
    if "'" not in kk:
        return False
    return first_vowel_is_full(kk)

suffix_pattern = re.compile(r'(ation|tion|sion|ution)$', re.IGNORECASE)

for row in ws.iter_rows(min_row=2):
    kk_cell = row[kk_col - 1]
    word_cell = row[word_col - 1]
    original = kk_cell.value
    word = word_cell.value

    if not original:
        continue

    kk = str(original)

    # Phase 1: Character substitutions
    kk = kk.replace('ˏ', 'ˌ')    # U+02CB wrong secondary stress -> U+02CC
    kk = kk.replace('ˋ', 'ˌ')    # U+02CB variant
    kk = kk.replace('͵', 'ˌ')    # U+0375 Greek lower numeral sign -> ˌ
    kk = kk.replace('∫', 'ʃ')    # U+222B integral sign -> U+0283 esh
    kk = kk.replace('∧', 'ʌ')    # U+2227 logical and -> U+028C turned v
    kk = kk.replace('ο', 'o')    # U+03BF Greek omicron -> Latin o
    kk = kk.replace('l̩', 'ḷ')   # IPA syllabic l -> KK syllabic l (dot below)
    kk = kk.replace('·', '')     # U+00B7 middle dot syllable separator -> remove
    kk = re.sub(r'\.(?!\d)', '', kk)   # remove period used as syllable separator (but not in numbers)
    kk = re.sub(r'\s+', '', kk)  # remove any whitespace
    kk = re.sub(r"'{2,}", "'", kk)  # collapse double '' -> '

    # Phase 2: Manual specific word fixes
    if word and word.lower() in manual:
        kk = manual[word.lower()]

    # Phase 3: Add missing ˌ to -ation/-tion/-sion/-ution words
    elif word and suffix_pattern.search(word.lower()):
        if needs_secondary_stress(word, kk):
            kk = 'ˌ' + kk

    if kk != original:
        fixes.append((word, original, kk))
        kk_cell.value = kk

print(f'\nTotal fixes: {len(fixes)}')
for w, old, new in fixes:
    print(f'  {w}: {old} -> {new}')

wb.save('data/word_list.xlsx')
print('\nSaved data/word_list.xlsx')

# Regenerate exports
wb2 = openpyxl.load_workbook('data/word_list.xlsx')
ws2 = wb2.active
headers2 = [cell.value for cell in ws2[1]]
rows = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    rows.append(dict(zip(headers2, row)))

with open('export/word_list.csv', 'w', newline='', encoding='utf-8-sig') as f:
    w = csv.DictWriter(f, fieldnames=headers2)
    w.writeheader()
    w.writerows(rows)

with open('export/word_list.json', 'w', encoding='utf-8') as f:
    json.dump(rows, f, ensure_ascii=False, separators=(',', ':'))

with open('export/word_list_pretty.json', 'w', encoding='utf-8') as f:
    json.dump(rows, f, ensure_ascii=False, indent=2)

print('Exports regenerated: csv, json, pretty.json')
