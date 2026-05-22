"""
為 word_list.xlsx 中未完成的單字生成例句
嚴格遵循新版 CLAUDE.md / grammar-patterns.md / data-schema.md / prompt-template.md 規則
使用 DeepSeek API（deepseek-chat）
可中斷續傳，每 50 筆自動存檔
"""

import os, sys, json, time, random, re
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import requests

# ─── 設定 ───────────────────────────────────────────────

EXCEL_PATH = "E:/DeepSeek/Voc_15000/data/word_list.xlsx"
LOG_PATH = "E:/DeepSeek/Voc_15000/logs/generation_log.txt"
CHECKPOINT_PATH = "E:/DeepSeek/Voc_15000/logs/checkpoint.json"
ENV_PATH = "E:/DeepSeek/Voc_15000/.env"
API_KEY = open(ENV_PATH).read().strip().split("=", 1)[1]
API_URL = "https://api.deepseek.com/v1/chat/completions"
MODEL = "deepseek-chat"
SAVE_INTERVAL = 50  # 每 N 筆存檔一次
MAX_RETRIES = 3

# ─── 句型分配 ───────────────────────────────────────────

A2_PATTERNS = ["simple", "compound", "adv_clause", "adj_clause"]
B1_PATTERNS = ["participial", "conditional_1", "conditional_2", "infinitive_purpose"]
B2_PATTERNS = ["inversion", "participial_perf"]

CEFR_MAP = {
    "simple": "A2", "compound": "A2", "adv_clause": "A2", "adj_clause": "A2",
    "participial": "B1", "conditional_1": "B1", "conditional_2": "B1",
    "infinitive_purpose": "B1",
    "participial_perf": "B1-B2", "inversion": "B1-B2",
}

def build_pattern_sequence(total_count):
    """
    依 60:30:10 比例分配 grammar_pattern，隨機打亂確保多樣性。
    """
    a2_count = round(total_count * 0.6)
    b1_count = round(total_count * 0.3)
    b2_count = total_count - a2_count - b1_count

    patterns = []
    # A2: 四種平均分配
    per_a2 = a2_count // 4
    for p in A2_PATTERNS:
        patterns.extend([p] * per_a2)
    remainder = a2_count - per_a2 * 4
    for i in range(remainder):
        patterns.append(A2_PATTERNS[i])

    # B1: participial 33.5%, conditional_1 16.5%, conditional_2 16.5%, infinitive_purpose 33.5%
    b1_pp = round(b1_count * 0.335)
    b1_c1 = round(b1_count * 0.165)
    b1_c2 = round(b1_count * 0.165)
    b1_ip = b1_count - b1_pp - b1_c1 - b1_c2
    patterns.extend([B1_PATTERNS[0]] * b1_pp)
    patterns.extend([B1_PATTERNS[1]] * b1_c1)
    patterns.extend([B1_PATTERNS[2]] * b1_c2)
    patterns.extend([B1_PATTERNS[3]] * b1_ip)

    # B1-B2: 各半
    per_b2 = b2_count // 2
    for p in B2_PATTERNS:
        patterns.extend([p] * per_b2)
    patterns.extend([B2_PATTERNS[0]] * (b2_count - per_b2 * 2))

    random.shuffle(patterns)
    return patterns


# ─── System Prompt ──────────────────────────────────────

SYSTEM_PROMPT = """You are an English teacher creating example sentences for Taiwanese high school students preparing for the CSAT (學測) exam.

Rules:
- Vocabulary: Taiwan MOE high school word list only.
- Length: 8-14 words per sentence.
- Natural tone; context relatable to Taiwanese high school students (school, family, technology, environment, social issues).
- Handle multi-POS words correctly.

KK phonetics (Taiwan style):
- In the `kk_phonetics` field: do NOT wrap in [ ] brackets; use ' for primary stress before the stressed vowel.
- In the `definition_zh` field: phonetics are ONLY included when the SAME word has DIFFERENT pronunciations across different POS (e.g. stress shift between noun and verb). MUST wrap in [ ] brackets.
- If the word has only ONE POS: fill `kk_phonetics` field; `definition_zh` contains NO phonetics.
- If multiple POS share the SAME phonetics: fill `kk_phonetics` field; `definition_zh` contains NO phonetics.
- If multiple POS have DIFFERENT phonetics: leave `kk_phonetics` EMPTY; annotate each POS in `definition_zh` as `(詞類) [音標] 譯1；譯2`.
- Phrasal verbs (片): NEVER include KK phonetics for this POS entry.
- Use KK symbols: ' for stress, ɛ for short-e, e for eɪ, ɝ/ɚ for r-colored, ɪ for short-i, θ/ð for th, ŋ for ng, ʃ for sh, ʒ for zh, ə for schwa.
- Example for "record" (名/動, different stress): kk_phonetics="", definition_zh="(名) ['rɛkɚd] 紀錄；唱片　(動) [rɪ'kɔrd] 記錄；錄音"
- Example for "light" (名/動/形, same pronunciation): kk_phonetics="laɪt", definition_zh="(名) 光；燈　(動) 點燃；照亮　(形) 輕的；淡的"
- Example for "abide by" (片): kk_phonetics="", definition_zh="(片) 遵守；遵循"

definition_zh: list up to 3 most common parts of speech of the target word, each with up to 2 most common Chinese translations. Use Chinese POS abbreviations: (名)(動)(形)(副)(介)(連)(代)(助)(片).

subject_type: use one of: common noun, proper name, gerund, infinitive, noun clause, abstract noun, compound noun.

Return ONLY a valid JSON object (no markdown, no code fences):
{
  "example_sentence": "...",
  "sentence_zh": "...",
  "cefr_level": "A2 / B1 / B1-B2",
  "grammar_pattern": "...",
  "word_count": number,
  "subject_type": "...",
  "kk_phonetics": "string or empty string",
  "definition_zh": "..."
}"""


def build_user_prompt(word, moe_level, grammar_pattern):
    """依 prompt-template.md 建立 user prompt"""
    lines = []
    lines.append(f'Generate ONE example sentence for the word: "{word}"')

    if moe_level is not None:
        lines.append(f"Word level: {moe_level}/6 (Taiwan MOE classification; 1 = most basic, 6 = most advanced; CSAT exam focuses on levels 1-4)")
        lines.append(f"For the supporting words (not the target word), prefer words at or below moe_level {moe_level}.")

    lines.append("")
    lines.append(f"Use the following grammar pattern: {grammar_pattern}")
    lines.append(f"(Options: simple, compound, adv_clause, adj_clause, participial, participial_perf, conditional_1, conditional_2, infinitive_purpose, inversion)")
    lines.append("")
    lines.append("CEFR mapping: simple/compound/adv_clause/adj_clause → A2; participial/conditional_1/conditional_2/infinitive_purpose → B1; participial_perf/inversion → B1-B2.")
    lines.append("")
    lines.append("IMPORTANT rules:")
    lines.append("- The target word MUST appear in the sentence in its correct part of speech.")
    lines.append("- Sentence must be 8-14 words (count actual words before punctuation).")
    lines.append("- Use diverse subjects. Avoid personal pronouns (he/she/they/we/I/you) as much as possible.")
    lines.append("- Use common English names, titles, professions, gerunds, noun clauses, or common nouns as subjects.")
    lines.append("- Context should be relatable to Taiwanese high school students.")
    lines.append("- For participial phrases: the participial subject must match the main clause subject.")
    lines.append("- For inversion: use only Never / Seldom / Not only…but also / Only after….")
    lines.append("- For conditional_2: If + past tense, would + base verb.")

    return "\n".join(lines)


# ─── API 呼叫 ───────────────────────────────────────────

def call_api(word, moe_level, grammar_pattern):
    """呼叫 DeepSeek API，回傳 parsed JSON dict"""
    user_prompt = build_user_prompt(word, moe_level, grammar_pattern)

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.post(
                API_URL,
                headers={"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"},
                json={
                    "model": MODEL,
                    "messages": [
                        {"role": "system", "content": SYSTEM_PROMPT},
                        {"role": "user", "content": user_prompt},
                    ],
                    "max_tokens": 500,
                    "temperature": 0.7,
                },
                timeout=60,
            )

            if resp.status_code != 200:
                log(f"  API error (attempt {attempt}): HTTP {resp.status_code} - {resp.text[:200]}")
                time.sleep(2)
                continue

            raw = resp.json()["choices"][0]["message"]["content"]

            # 提取 JSON（處理可能的 markdown code fence）
            json_str = raw.strip()
            if "```json" in json_str:
                json_str = json_str.split("```json")[1].split("```")[0]
            elif "```" in json_str:
                json_str = json_str.split("```")[1].split("```")[0]

            data = json.loads(json_str.strip())

            # 基本驗證
            required = ["example_sentence", "sentence_zh", "cefr_level",
                        "grammar_pattern", "word_count", "subject_type",
                        "kk_phonetics", "definition_zh"]
            for key in required:
                if key not in data:
                    raise ValueError(f"Missing key: {key}")

            # 驗證字數
            actual_wc = len(data["example_sentence"]
                          .replace(".","").replace(",","").replace("!","")
                          .replace("?","").replace("'","").replace('"',"")
                          .split())
            if not (8 <= actual_wc <= 14):
                raise ValueError(f"Word count {actual_wc} not in 8-14 range")

            data["_actual_word_count"] = actual_wc
            return data

        except (json.JSONDecodeError, ValueError, requests.RequestException) as e:
            log(f"  Attempt {attempt} failed: {e}")
            if attempt < MAX_RETRIES:
                time.sleep(3)

    return None


# ─── 寫入 Excel ────────────────────────────────────────

DFONT = Font(name="Arial Unicode MS", size=10)
DALIGN = Alignment(vertical="center", wrap_text=True)
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def write_row(ws, row_num, data, grammar_pattern, cefr_level):
    """將 API 回傳結果寫入 Excel 指定列"""
    wc = data.get("_actual_word_count",
                  len(data["example_sentence"]
                      .replace(".","").replace(",","").replace("!","")
                      .replace("?","").replace("'","").replace('"',"")
                      .split()))
    kk = data.get("kk_phonetics", "")
    def_zh = data.get("definition_zh", "")
    subject_type = data.get("subject_type", "common noun")

    # 欄位順序：id(1), word(2), moe_level(3), kk_phonetics(4), definition_zh(5),
    #          example_sentence(6), sentence_zh(7), cefr_level(8), grammar_pattern(9),
    #          word_count(10), subject_type(11), source(12), reviewed(13), notes(14)
    # 注意：id 和 word 已有值，不覆寫；moe_level 也已有值不覆寫
    col_values = {
        4: kk,          # kk_phonetics
        5: def_zh,      # definition_zh
        6: data["example_sentence"],  # example_sentence
        7: data["sentence_zh"],       # sentence_zh
        8: cefr_level,  # cefr_level
        9: grammar_pattern,  # grammar_pattern
        10: wc,         # word_count
        11: subject_type,  # subject_type
        12: "Claude AI",  # source
        13: "FALSE",     # reviewed
        14: "",          # notes
    }

    for col_num, val in col_values.items():
        cell = ws.cell(row=row_num, column=col_num, value=val)
        cell.font = DFONT
        cell.alignment = DALIGN
        cell.border = BORDER


# ─── 日誌 ──────────────────────────────────────────────

def log(msg):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] {msg}"
    print(line)
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(line + "\n")


# ─── 主流程 ─────────────────────────────────────────────

def main():
    log("=== 開始生成例句 ===")

    # 載入 Excel
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    total_rows = ws.max_row

    # 找出所有未完成的列（example_sentence 為空）
    pending = []
    for row in range(2, total_rows + 1):
        example = ws.cell(row, 6).value
        if not example or not str(example).strip():
            rid = ws.cell(row, 1).value       # id
            word = ws.cell(row, 2).value       # word
            moe = ws.cell(row, 3).value        # moe_level
            pending.append({
                "row": row,
                "id": rid,
                "word": str(word).strip().lower() if word else "",
                "moe_level": int(moe) if moe is not None else None,
            })

    total_pending = len(pending)
    log(f"找到 {total_pending} 個未完成單字")

    if total_pending == 0:
        log("沒有未完成的單字，結束。")
        return

    # 分配句型
    patterns = build_pattern_sequence(total_pending)
    log(f"句型分配完成（共 {len(patterns)} 個）")

    # 統計用
    stats = {"success": 0, "fail": 0, "skip": 0}

    for idx, (item, gp) in enumerate(zip(pending, patterns)):
        rid = item["id"]
        word = item["word"]
        moe = item["moe_level"]
        row_num = item["row"]
        cefr = CEFR_MAP[gp]

        log(f"[{idx+1}/{total_pending}] id={rid}, word={word}, moe={moe}, grammar={gp}")

        # 再次確認該列是否已填（支援續傳）
        existing = ws.cell(row_num, 6).value
        if existing and str(existing).strip():
            log(f"  已存在例句，跳過")
            stats["skip"] += 1
            continue

        # 呼叫 API
        result = call_api(word, moe, gp)

        if result is None:
            log(f"  ❌ API 失敗（已重試 {MAX_RETRIES} 次），跳過")
            stats["fail"] += 1
            continue

        # 寫入 Excel
        write_row(ws, row_num, result, gp, cefr)
        stats["success"] += 1

        log(f"  ✅ {result['example_sentence'][:60]}...")

        # 定期存檔
        if stats["success"] % SAVE_INTERVAL == 0:
            wb.save(EXCEL_PATH)
            log(f"  💾 已存檔（累計成功 {stats['success']} 筆）")

        # API 流量控制
        time.sleep(0.5)

    # 最終存檔
    wb.save(EXCEL_PATH)
    log("=== 完成 ===")
    log(f"成功: {stats['success']}, 失敗: {stats['fail']}, 跳過: {stats['skip']}")
    log(f"總處理: {stats['success'] + stats['fail'] + stats['skip']}")


if __name__ == "__main__":
    main()
