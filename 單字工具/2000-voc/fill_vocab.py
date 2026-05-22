#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl

# Comprehensive vocabulary data
vocab_data = {
    'able': ('[ˈeɪbəl]', '(形) 能...的'),
    'about': ('[əˈbaʊt]', '(介) 關於'),
    'above': ('[əˈbʌv]', '(介) 在...上方'),
    'accept': ('[əkˈsɛpt]', '(動) 接受'),
    'accident': ('[ˈæksɪdənt]', '(名) 意外；事故'),
    'across': ('[əˈkrɔs]', '(介) 穿過'),
    'act': ('[ækt]', '(動) 行動；表演'),
    'action': ('[ˈækʃən]', '(名) 行動'),
    'active': ('[ˈæktɪv]', '(形) 活躍的'),
    'activity': ('[ækˈtɪvəti]', '(名) 活動'),
    'add': ('[æd]', '(動) 添加；加上'),
    'address': ('[ˈædrɛs]', '(名) 地址'),
    'afraid': ('[əˈfreɪd]', '(形) 害怕的'),
    'after': ('[ˈæftɚ]', '(介) 在...之後'),
    'again': ('[əˈɡɛn]', '(副) 再；又'),
    'age': ('[eɪdʒ]', '(名) 年齡'),
    'agree': ('[əˈɡri]', '(動) 同意'),
    'ahead': ('[əˈhɛd]', '(副) 在前面'),
    'air': ('[ɛr]', '(名) 空氣'),
    'all': ('[ɔl]', '(代) 全部'),
    'allow': ('[əˈlaʊ]', '(動) 允許'),
    'almost': ('[ˈɔlmoʊst]', '(副) 幾乎'),
    'alone': ('[əˈloʊn]', '(形) 獨自的'),
    'along': ('[əˈlɔŋ]', '(介) 沿著'),
    'also': ('[ˈɔlsoʊ]', '(副) 也'),
    'always': ('[ˈɔlweɪz]', '(副) 總是'),
    'among': ('[əˈmʌŋ]', '(介) 在...之間'),
    'and': ('[ənd]', '(連) 和'),
    'animal': ('[ˈænɪməl]', '(名) 動物'),
    'another': ('[əˈnʌðɚ]', '(代) 另一個'),
    'answer': ('[ˈænsɚ]', '(名/動) 答案；回答'),
    'any': ('[ˈɛni]', '(代) 任何'),
    'anyone': ('[ˈɛniˌwʌn]', '(代) 任何人'),
    'anything': ('[ˈɛniˌθɪŋ]', '(代) 任何東西'),
    'apartment': ('[əˈpɑrtmənt]', '(名) 公寓'),
    'appear': ('[əˈpɪr]', '(動) 出現'),
    'apple': ('[ˈæpəl]', '(名) 蘋果'),
    'area': ('[ˈɛriə]', '(名) 面積；地區'),
    'argue': ('[ˈɑrɡju]', '(動) 爭論'),
    'arm': ('[ɑrm]', '(名) 手臂'),
    'around': ('[əˈraʊnd]', '(介) 在...周圍'),
    'arrive': ('[əˈraɪv]', '(動) 到達'),
    'art': ('[ɑrt]', '(名) 藝術'),
    'artist': ('[ˈɑrtɪst]', '(名) 藝術家'),
    'as': ('[æz]', '(連) 當...時；因為'),
    'ask': ('[æsk]', '(動) 問；要求'),
    'asleep': ('[əˈslip]', '(形) 睡著的'),
    'attack': ('[əˈtæk]', '(動/名) 攻擊'),
    'attention': ('[əˈtɛnʃən]', '(名) 注意'),
    'available': ('[əˈveɪləbəl]', '(形) 可取得的'),
    'average': ('[ˈævərɪdʒ]', '(形/名) 平均'),
    'avoid': ('[əˈvɔɪd]', '(動) 避免'),
    'awake': ('[əˈweɪk]', '(形/動) 醒著；叫醒'),
    'away': ('[əˈweɪ]', '(副) 離開'),
}

try:
    wb = openpyxl.load_workbook('2000voc.xlsx')
    ws = wb.active
    
    count = 0
    for row in range(2, ws.max_row + 1):
        word = ws[f'A{row}'].value
        if word and word.lower() in vocab_data:
            if not ws[f'B{row}'].value:
                phonetic, pos1 = vocab_data[word.lower()]
                ws[f'B{row}'] = phonetic
                ws[f'C{row}'] = pos1
                count += 1
    
    wb.save('2000voc.xlsx')
    print('Success: {} words updated'.format(count))
    
except Exception as e:
    print('Error: {}'.format(str(e)))
