#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl

# Extended vocabulary data for 100+ high-frequency words
vocab_data = {
    'baby': ('[ˈbeɪbi]', '(名) 嬰兒'),
    'back': ('[bæk]', '(名) 背部'),
    'bad': ('[bæd]', '(形) 壞的'),
    'bag': ('[bæɡ]', '(名) 袋子'),
    'ball': ('[bɔl]', '(名) 球'),
    'band': ('[bænd]', '(名) 樂隊'),
    'bank': ('[bæŋk]', '(名) 銀行；河岸'),
    'basketball': ('[ˈbæskətˌbɔl]', '(名) 籃球'),
    'bathroom': ('[ˈbæθˌrum]', '(名) 浴室'),
    'beach': ('[bitʃ]', '(名) 海灘'),
    'bear': ('[bɛr]', '(名) 熊'),
    'beautiful': ('[ˈbjutəfəl]', '(形) 美麗的'),
    'because': ('[bɪˈkɔz]', '(連) 因為'),
    'become': ('[bɪˈkʌm]', '(動) 變成'),
    'bed': ('[bɛd]', '(名) 床'),
    'before': ('[bɪˈfɔr]', '(介) 在...之前'),
    'begin': ('[bɪˈɡɪn]', '(動) 開始'),
    'believe': ('[bɪˈliv]', '(動) 相信'),
    'belong': ('[bɪˈlɔŋ]', '(動) 屬於'),
    'below': ('[bɪˈloʊ]', '(介) 在...下方'),
    'beside': ('[bɪˈsaɪd]', '(介) 在...旁邊'),
    'best': ('[bɛst]', '(形) 最好的'),
    'better': ('[ˈbɛtɚ]', '(形) 更好的'),
    'between': ('[bɪˈtwin]', '(介) 在...之間'),
    'beyond': ('[bɪˈjɑnd]', '(介) 在...之外'),
    'big': ('[bɪɡ]', '(形) 大的'),
    'bird': ('[bɜrd]', '(名) 鳥'),
    'birth': ('[bɜrθ]', '(名) 出生'),
    'birthday': ('[ˈbɜrθdeɪ]', '(名) 生日'),
    'bit': ('[bɪt]', '(名) 一點'),
    'black': ('[blæk]', '(形) 黑色的'),
    'blue': ('[blu]', '(形) 藍色的'),
    'board': ('[bɔrd]', '(名) 板；委員會'),
    'boat': ('[boʊt]', '(名) 船'),
    'body': ('[ˈbɑdi]', '(名) 身體'),
    'book': ('[bʊk]', '(名) 書'),
    'bring': ('[brɪŋ]', '(動) 帶來'),
    'brother': ('[ˈbrʌðɚ]', '(名) 哥哥；弟弟'),
    'bus': ('[bʌs]', '(名) 公車'),
    'business': ('[ˈbɪznəs]', '(名) 業務'),
    'but': ('[bət]', '(連) 但是'),
    'buy': ('[baɪ]', '(動) 買'),
    'by': ('[baɪ]', '(介) 被；在...旁邊'),
    'call': ('[kɔl]', '(動) 叫；打電話'),
    'can': ('[kæn]', '(助) 能；可以'),
    'card': ('[kɑrd]', '(名) 卡片'),
    'care': ('[kɛr]', '(動/名) 照顧；關心'),
    'cat': ('[kæt]', '(名) 貓'),
    'catch': ('[kætʃ]', '(動) 抓住'),
    'center': ('[ˈsɛntɚ]', '(名) 中心'),
    'certainly': ('[ˈsɜrtənli]', '(副) 確實'),
    'change': ('[tʃeɪndʒ]', '(動/名) 改變'),
    'character': ('[ˈkɛrəktɚ]', '(名) 字元；性格'),
    'cheap': ('[tʃip]', '(形) 便宜的'),
    'check': ('[tʃɛk]', '(動) 檢查'),
    'cheese': ('[tʃiz]', '(名) 起司'),
    'chicken': ('[ˈtʃɪkən]', '(名) 雞'),
    'child': ('[tʃaɪld]', '(名) 孩子'),
    'choice': ('[tʃɔɪs]', '(名) 選擇'),
    'choose': ('[tʃuz]', '(動) 選擇'),
    'church': ('[tʃɜrtʃ]', '(名) 教堂'),
    'circle': ('[ˈsɜrkəl]', '(名) 圓圈'),
    'city': ('[ˈsɪti]', '(名) 城市'),
    'class': ('[klæs]', '(名) 班級'),
    'clean': ('[klin]', '(動/形) 清潔；乾淨'),
    'clear': ('[klɪr]', '(形) 清楚的'),
    'close': ('[kloʊz]', '(動) 關閉'),
    'club': ('[klʌb]', '(名) 俱樂部'),
    'coffee': ('[ˈkɔfi]', '(名) 咖啡'),
    'cold': ('[koʊld]', '(形/名) 冷的；感冒'),
    'color': ('[ˈkʌlɚ]', '(名) 顏色'),
    'come': ('[kʌm]', '(動) 來'),
    'common': ('[ˈkɑmən]', '(形) 普通的'),
    'company': ('[ˈkʌmpəni]', '(名) 公司'),
    'complete': ('[kəmˈplit]', '(形/動) 完整；完成'),
    'computer': ('[kəmˈpjutɚ]', '(名) 電腦'),
    'concern': ('[kənˈsɜrn]', '(動/名) 關心；問題'),
    'condition': ('[kənˈdɪʃən]', '(名) 條件'),
    'consider': ('[kənˈsɪdɚ]', '(動) 考慮'),
    'continue': ('[kənˈtɪnju]', '(動) 繼續'),
    'control': ('[kənˈtroʊl]', '(動/名) 控制'),
    'conversation': ('[ˌkɑnvɚˈseɪʃən]', '(名) 對話'),
    'cook': ('[kʊk]', '(動/名) 烹飪；廚師'),
    'cool': ('[kul]', '(形) 涼爽的'),
    'corner': ('[ˈkɔrnɚ]', '(名) 角落'),
    'correct': ('[kəˈrɛkt]', '(形/動) 正確；改正'),
    'cost': ('[kɔst]', '(動/名) 花費；成本'),
    'cotton': ('[ˈkɑtən]', '(名) 棉花'),
    'could': ('[kʊd]', '(助) 可能'),
    'country': ('[ˈkʌntri]', '(名) 國家'),
    'course': ('[kɔrs]', '(名) 課程'),
    'cousin': ('[ˈkʌzən]', '(名) 堂（表）兄弟姐妹'),
    'cover': ('[ˈkʌvɚ]', '(動/名) 覆蓋'),
    'cow': ('[kaʊ]', '(名) 牛'),
    'crazy': ('[ˈkreɪzi]', '(形) 瘋狂的'),
    'cream': ('[krim]', '(名) 奶油'),
    'create': ('[kriˈeɪt]', '(動) 創造'),
    'cross': ('[krɔs]', '(動/形) 穿過；生氣'),
    'crowd': ('[kraʊd]', '(名) 人群'),
    'cry': ('[kraɪ]', '(動) 哭；喊'),
    'cup': ('[kʌp]', '(名) 杯子'),
    'culture': ('[ˈkʌltʃɚ]', '(名) 文化'),
    'curious': ('[ˈkjʊriəs]', '(形) 好奇的'),
    'cut': ('[kʌt]', '(動) 切割'),
}

try:
    wb = openpyxl.load_workbook('2000voc.xlsx')
    ws = wb.active
    
    count = 0
    for row in range(2, ws.max_row + 1):
        word = ws[f'A{row}'].value
        if word and word.lower() in vocab_data:
            if not ws[f'B{row}'].value:
                phonetic, pos1 = vocab_data[word.lower()]
                ws[f'B{row}'] = phonetic
                ws[f'C{row}'] = pos1
                count += 1
    
    wb.save('2000voc.xlsx')
    print('Success: {} more words updated'.format(count))
    
except Exception as e:
    print('Error: {}'.format(str(e)))
