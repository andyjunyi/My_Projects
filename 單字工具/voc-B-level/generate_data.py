"""
generate_data.py
────────────────
從「國中2000單高頻單字.xlsx」提取資料，輸出 vocab_data.js。
網頁 index.html 直接以 <script src> 載入，不需要 Web Server。

用法：
    python generate_data.py
    # 或指定不同的 Excel 檔案：
    python generate_data.py --excel 其他檔案.xlsx

新增/修改單字：
    直接在 Excel 的「2000」工作表新增/編輯列，
    在「考題高頻單字935_對應2000單排名」工作表補上頻率資料（可選），
    然後重新執行此腳本即可。
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("請先安裝 openpyxl：pip install openpyxl")


# ─────────────────────────────────────────────────
# 音節數估算（由 KK 音標）
# ─────────────────────────────────────────────────
# KK 複合母音（diphthong）→ 1 音節
_DIPHTHONGS = ['aɪ', 'aʊ', 'ɔɪ', 'eɪ', 'oʊ', 'ɪr', 'ɛr', 'ʊr', 'ɑr']
# KK 單母音（monophthong）→ 1 音節
_MONOPHTHONGS = set('æɛɑɔʊʌəɪɝɚiueoaɔ')

def count_syllables(phonetic: str) -> int:
    """計算 KK 音標的音節數（以母音核心為準）。"""
    if not phonetic:
        return 0
    # 移除重音符號、長音符號等修飾符
    clean = phonetic.replace('ˈ', '').replace('ˌ', '').replace('ː', '')
    count = 0
    i = 0
    while i < len(clean):
        matched = False
        # 先嘗試雙母音
        for d in _DIPHTHONGS:
            if clean[i:i+len(d)] == d:
                count += 1
                i += len(d)
                matched = True
                break
        if not matched:
            if clean[i] in _MONOPHTHONGS:
                count += 1
            i += 1
    return max(1, count)


# ─────────────────────────────────────────────────
# 主程式
# ─────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='產生 vocab_data.js')
    parser.add_argument('--excel', default='國中2000單高頻單字.xlsx',
                        help='Excel 來源檔（預設：國中2000單高頻單字.xlsx）')
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        sys.exit(f'找不到 Excel 檔案：{excel_path}')

    print(f'讀取：{excel_path}')
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # ── 1. 主詞庫「2000」工作表 ──────────────────────
    vocab: dict = {}
    if '2000' in wb.sheetnames:
        ws = wb['2000']
        for row in ws.iter_rows(min_row=2, values_only=True):
            word = row[0]
            if not word or not isinstance(word, str):
                continue
            key = word.strip().lower()
            vocab[key] = {
                'word':       word.strip(),
                'phonetic':   str(row[1]).strip() if row[1] else '',
                'chinese':    str(row[2]).strip() if row[2] else '',
                'example':    str(row[3]).strip() if row[3] else '',
                'example_zh': str(row[4]).strip() if row[4] else '',
                'pos':        str(row[5]).strip() if row[5] else '',
            }
        print(f'  2000 工作表：{len(vocab)} 個單字')
    else:
        print('警告：找不到「2000」工作表，跳過主詞庫。')

    # ── 2. 考題頻率「考題高頻單字935_對應2000單排名」 ─
    freq: dict = {}
    FREQ_SHEET = '考題高頻單字935_對應2000單排名'
    if FREQ_SHEET in wb.sheetnames:
        ws = wb[FREQ_SHEET]
        CHECKMARK = {'✓', 'V', 'v', '1', 'True', 'true', 'TRUE', 'yes', 'Yes'}
        for row in ws.iter_rows(min_row=2, values_only=True):
            rank = row[0]
            word = row[1]
            if not word or not isinstance(word, str):
                continue
            if not isinstance(rank, (int, float)):
                continue
            key = word.strip().lower()
            freq[key] = {
                'rank_935':           int(rank),
                'frequency':          int(row[6])   if isinstance(row[6], (int,float)) else 0,
                'years_count':        int(row[7])   if isinstance(row[7], (int,float)) else 0,
                'years_list':         str(row[8]).strip() if row[8] else '',
                'freq_percent':       float(row[9]) if isinstance(row[9], (int,float)) else 0.0,
                'cumulative_percent': float(row[10])if isinstance(row[10],(int,float))else 0.0,
                'cefr':               str(row[11]).strip() if row[11] else '',
                'level':              str(row[12]).strip() if row[12] else '',
                'topic':              str(row[14]).strip() if row[14] else '',
                'core300':            str(row[15]).strip() in CHECKMARK if row[15] else False,
                'consolidate450':     str(row[16]).strip() in CHECKMARK if row[16] else False,
            }
        print(f'  {FREQ_SHEET}：{len(freq)} 個高頻單字')
    else:
        print(f'警告：找不到「{FREQ_SHEET}」工作表，不含頻率資料。')

    # ── 3. 合併 ─────────────────────────────────────
    words = []
    for key, v in vocab.items():
        entry = dict(v)
        if key in freq:
            entry.update(freq[key])
        else:
            entry.update({
                'rank_935': None, 'frequency': 0, 'years_count': 0,
                'years_list': '', 'freq_percent': 0.0, 'cumulative_percent': 0.0,
                'cefr': '', 'level': '', 'topic': '',
                'core300': False, 'consolidate450': False,
            })
        entry['syllable_count'] = count_syllables(entry['phonetic'])
        words.append(entry)

    # 依字母排序
    words.sort(key=lambda x: x['word'].lower())

    # ── 4. 輸出 vocab_data.js（可直接由 file:// 載入）──
    out_path = Path('vocab_data.js')
    data = {'words': words, 'total': len(words)}
    json_str = json.dumps(data, ensure_ascii=False)
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write('/* 自動產生，請勿手動編輯 — 更新請重執行 generate_data.py */\n')
        f.write(f'window.VOCAB_DATA = {json_str};\n')

    # 同時保留 JSON 備份
    json_path = Path('vocab_data.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f'輸出：{out_path}（共 {len(words)} 個單字，'
          f'{out_path.stat().st_size/1024:.1f} KB）')

    # ── 5. 輸出 exam_data.js（試題索引）────────────────
    generate_exam_data()
    print('完成！重新整理瀏覽器即可看到更新後的資料。')


# ─────────────────────────────────────────────────
# 試題資料提取
# ─────────────────────────────────────────────────
def generate_exam_data():
    import html as html_mod

    SKIP = {'index.html', 'guide.html'}
    html_files = sorted(
        p for p in Path('.').glob('*.html') if p.name not in SKIP
    )
    if not html_files:
        print('未找到試題 HTML 檔案，略過 exam_data.js')
        return

    def strip_tags(s):
        return re.sub(r'<[^>]+>', '', s)

    def get_text(s):
        return html_mod.unescape(strip_tags(s)).strip() if s else ''

    all_exams = []

    for fpath in html_files:
        content = fpath.read_text(encoding='utf-8')

        title_m = re.search(r'<title>(.*?)</title>', content)
        title = html_mod.unescape(title_m.group(1)) if title_m else fpath.name

        # Answer key
        ans_m = re.search(r'const ANSWERS\s*=\s*\{([^}]+)\}', content)
        answers: dict = {}
        if ans_m:
            for m in re.finditer(r"q(\d+)\s*:\s*'([A-D])'", ans_m.group(1)):
                answers[m.group(1)] = m.group(2)

        # Each q-card block
        # Match up to the closing </div> of .explanation then two more </div>
        card_pat = re.compile(
            r'<div class="q-card" id="qcard-(\d+)">(.*?)</div>\s*\n?\s*</div>',
            re.DOTALL
        )

        questions = []
        for qnum, body in card_pat.findall(content):
            # Question text – keep blank marker as [____]
            qt_m = re.search(r'class="q-text">(.*?)</p>', body, re.DOTALL)
            if qt_m:
                qt_raw = qt_m.group(1)
                # Replace blank span with visible placeholder for plain-text search
                qt_plain = get_text(qt_raw.replace(
                    '<span class="blank">&nbsp;</span>', '____'))
                # Keep HTML for display (with blank styled)
                qt_html = qt_raw.strip()
            else:
                qt_plain = qt_html = ''

            # Options
            opts_raw = re.findall(
                r'<label class="opt-label"><input[^>]+value="([A-D])">\s*(.*?)</label>',
                body, re.DOTALL
            )
            options = {v: get_text(t) for v, t in opts_raw}

            # Full sentence & Chinese translation from explanation
            full_m = re.search(r'class="full-sent">(.*?)</div>', body, re.DOTALL)
            en_full = zh_full = ''
            if full_m:
                raw = get_text(full_m.group(1))
                parts = re.split(r'中文翻譯[：:]', raw)
                en_full = re.sub(r'完整句子[：:]', '', parts[0]).strip()
                zh_full = parts[1].strip() if len(parts) > 1 else ''

            # Search index: all searchable text merged (lowercase)
            search_text = ' '.join(filter(None, [
                qt_plain,
                ' '.join(options.values()),
                en_full,
            ])).lower()

            questions.append({
                'num':         int(qnum),
                'q_text':      qt_plain,
                'q_html':      qt_html,
                'options':     options,
                'answer':      answers.get(qnum, ''),
                'en_full':     en_full,
                'zh_full':     zh_full,
                'search_text': search_text,
            })

        questions.sort(key=lambda q: q['num'])
        all_exams.append({
            'file':      fpath.name,
            'title':     title,
            'questions': questions,
        })
        print(f'  {fpath.name}：{len(questions)} 題')

    out = Path('exam_data.js')
    js_str = json.dumps(all_exams, ensure_ascii=False)
    with open(out, 'w', encoding='utf-8') as f:
        f.write('/* 自動產生，請勿手動編輯 — 更新請重執行 generate_data.py */\n')
        f.write(f'window.EXAM_DATA = {js_str};\n')

    total_q = sum(len(e['questions']) for e in all_exams)
    print(f'輸出：{out}（{len(all_exams)} 份試題，共 {total_q} 題，'
          f'{out.stat().st_size/1024:.1f} KB）')


if __name__ == '__main__':
    main()
