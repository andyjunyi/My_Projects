#!/usr/bin/env python3
"""
KK 音標 / 資料同步腳本
從修改記錄檔 (tsv) 讀取變更 → 更新 word_list.xlsx → 重新產生所有輸出檔

輸入格式（Tab 分隔）：
  單字\t欄位\t原值\t新值

範例：
  conference\tk\t'kɑnfɚəns\t'kɑnfərəns
  ability\tz\t(名) 能力；才能\t(名) 能力；技能
  abandon\tl\t1\t2
  ability\tc\tA2\tB1

欄位代號：
  k  = kk_phonetics
  z  = definition_zh (中譯)
  e  = example_sentence (例句)
  ez = sentence_zh (例句中譯)
  l  = moe_level (級別)
  c  = cefr_level
"""

import sys, os, json

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(BASE, "word_list.xlsx")
JSON_PATH = os.path.join(BASE, "word_data.json")
VOC_15000_DIR = os.path.join(os.path.dirname(os.path.dirname(BASE)), "Voc_15000")
EXPORT_DIR = os.path.join(VOC_15000_DIR, "export")

FIELD_MAP = {
    'k': 'kk_phonetics',
    'z': 'definition_zh',
    'e': 'example_sentence',
    'ez': 'sentence_zh',
    'l': 'moe_level',
    'c': 'cefr_level',
}

def read_changes():
    """從 stdin 讀取修改記錄"""
    changes = []
    for line in sys.stdin:
        line = line.strip()
        if not line or line.startswith('#'):
            continue
        parts = line.split('\t')
        if len(parts) >= 4:
            word = parts[0].strip()
            field_code = parts[1].strip()
            new_val = parts[3].strip()
            changes.append((word, field_code, new_val))
    return changes

def update_xlsx(changes):
    """更新 Excel"""
    from openpyxl import load_workbook

    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    headers = {}
    for i, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1))):
        if cell.value:
            headers[cell.value] = i

    updated = 0
    not_found = []
    skipped = []

    for word, field_code, new_val in changes:
        if field_code not in FIELD_MAP:
            skipped.append((word, f'未知欄位代號: {field_code}'))
            continue

        xlsx_field = FIELD_MAP[field_code]
        if xlsx_field not in headers:
            skipped.append((word, f'Excel 無欄位: {xlsx_field}'))
            continue

        col_idx = headers[xlsx_field]
        found = False
        for row in ws.iter_rows(min_row=2):
            cell_word = row[headers.get('word', 1)]
            if cell_word.value and str(cell_word.value).strip() == word:
                row[col_idx].value = new_val
                updated += 1
                found = True
                break
        if not found:
            not_found.append(word)

    wb.save(XLSX_PATH)
    return updated, not_found, skipped

def regenerate_json():
    """從 Excel 重新產生 word_data.json"""
    from openpyxl import load_workbook

    wb = load_workbook(XLSX_PATH, read_only=True)
    ws = wb.active

    words = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        word = row[1]
        if word is None:
            continue
        item = {
            "w": str(word) if word else "",
            "k": str(row[3]) if row[3] else "",
            "z": str(row[4]) if row[4] else "",
            "e": str(row[5]) if row[5] else "",
            "ez": str(row[6]) if row[6] else "",
            "l": str(row[2]) if row[2] else "",
            "c": str(row[7]) if row[7] else "",
        }
        words.append(item)

    wb.close()

    with open(JSON_PATH, 'w', encoding='utf-8') as f:
        json.dump(words, f, ensure_ascii=False, separators=(',', ':'))

    return len(words)

def sync_export():
    """同步到 Voc_15000/export/ 目錄的完整 JSON/CSV"""
    if not os.path.exists(EXPORT_DIR):
        return False

    from openpyxl import load_workbook
    import csv

    wb = load_workbook(XLSX_PATH, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [row for row in ws.iter_rows(min_row=2, values_only=True)]

    wb.close()

    full_json = []
    for row in rows:
        item = {}
        for i, h in enumerate(headers):
            item[h] = str(row[i]) if row[i] else ''
        full_json.append(item)

    with open(os.path.join(EXPORT_DIR, 'word_list.json'), 'w', encoding='utf-8') as f:
        json.dump(full_json, f, ensure_ascii=False)

    with open(os.path.join(EXPORT_DIR, 'word_list_pretty.json'), 'w', encoding='utf-8') as f:
        json.dump(full_json, f, ensure_ascii=False, indent=2)

    with open(os.path.join(EXPORT_DIR, 'word_list.csv'), 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)

    return True

def main():
    changes = read_changes()

    if not changes:
        print("❌ 沒有讀取到修改記錄")
        print("用法：python3 sync_kk.py < changes.tsv")
        sys.exit(1)

    print(f"📋 讀取到 {len(changes)} 筆修改")
    for w, f, v in changes:
        fname = FIELD_MAP.get(f, f)
        print(f"   {w:20s} [{fname}] → {v[:50]}")

    # Step 1: Excel
    print(f"\n🔧 更新 {os.path.basename(XLSX_PATH)}...")
    updated, not_found, skipped = update_xlsx(changes)
    print(f"   ✅ 已更新 {updated} 筆")
    if not_found:
        print(f"   ⚠️ 未找到：{', '.join(not_found[:10])}")
        if len(not_found) > 10:
            print(f"      ...還有 {len(not_found)-10} 筆")
    if skipped:
        for w, reason in skipped:
            print(f"   ⚠️ 跳過 {w}: {reason}")

    # Step 2: JSON
    print(f"\n🔄 重新產生 {os.path.basename(JSON_PATH)}...")
    count = regenerate_json()
    print(f"   ✅ 已輸出 {count} 筆")

    # Step 3: Export
    print(f"\n🔄 同步到 export/ 目錄...")
    if sync_export():
        print(f"   ✅ word_list.json / csv 已更新")
    else:
        print(f"   ⚪ export/ 目錄不存在，跳過")

    print(f"\n✅ 全部完成！")

if __name__ == '__main__':
    main()
