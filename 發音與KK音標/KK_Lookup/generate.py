import openpyxl, json, sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('國中935單.xlsx')

ws935 = wb['935']
words_935 = set()
for row in ws935.iter_rows(min_row=2, values_only=True):
    if row[1]:
        words_935.add(str(row[1]).strip().lower())

ws2000 = wb['2000']
db = {}
for row in ws2000.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    word = str(row[0]).strip()
    kk    = str(row[1]).strip() if row[1] else ''
    trans = str(row[2]).strip() if row[2] else ''
    pos_zh = str(row[5]).strip() if row[5] else ''
    db[word.lower()] = {
        'w': word, 'kk': kk, 't': trans, 'p': pos_zh,
        'g': 1 if word.lower() in words_935 else 0
    }

db_json = json.dumps(db, ensure_ascii=False)

# ---------------------------------------------------------------------------
# HTML template — uses __DB_JSON__ placeholder for the embedded vocabulary
# JS curly-brace literals are left as-is (raw string, no f-string)
# ---------------------------------------------------------------------------
HTML = r"""<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>國中英文單字查詢</title>
<style>
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: "Microsoft JhengHei", "PingFang TC", Arial, sans-serif;
    background: #f4f6f8; color: #222; padding: 30px 16px;
  }
  .container { max-width: 680px; margin: 0 auto; }
  h1 { font-size: 1.5rem; margin-bottom: 20px; color: #1a3a5c; }

  /* Search */
  .search-row { display: flex; gap: 8px; margin-bottom: 20px; }
  #searchInput {
    flex: 1; padding: 10px 14px; font-size: 1rem;
    border: 2px solid #c5d3e0; border-radius: 6px; outline: none;
  }
  #searchInput:focus { border-color: #2e7dd1; }
  .btn-primary {
    padding: 10px 22px; font-size: 1rem; cursor: pointer;
    background: #2e7dd1; color: #fff; border: none; border-radius: 6px;
  }
  .btn-primary:hover { background: #1b5fa8; }

  /* Result card */
  .result-card {
    background: #fff; border-radius: 8px;
    border: 1px solid #d0dce8; padding: 18px 20px;
    margin-bottom: 16px; position: relative;
  }
  .result-card.not-found  { border-color: #e0c0c0; background: #fff8f8; color: #a00; }
  .result-card.ai-pending { border-color: #e0d080; background: #fffdf0; }
  .result-card.loading    { border-color: #b0c8e8; background: #f0f6ff; color: #446; }
  .result-text {
    font-family: "Courier New", Consolas, monospace;
    font-size: 1.05rem; white-space: pre-wrap; line-height: 1.7; color: #111;
    word-break: break-word;
  }
  .copy-btn {
    position: absolute; top: 14px; right: 14px;
    padding: 5px 14px; font-size: 0.85rem; cursor: pointer;
    background: #eef3f9; border: 1px solid #b0c4dc; border-radius: 5px;
    color: #2e7dd1; font-weight: bold;
  }
  .copy-btn:hover { background: #d8e8f5; }
  .source-tag { margin-top: 8px; font-size: 0.78rem; color: #888; }
  .tag-935      { color: #1a7a3c; font-weight: bold; }
  .tag-supp     { color: #8a6000; }
  .tag-user     { color: #5a1a8a; }
  .tag-ai       { color: #a05000; font-weight: bold; }
  .tag-confirmed{ color: #1a7a3c; font-weight: bold; }

  /* Confirm row */
  .confirm-row { margin-top: 12px; display: flex; gap: 10px; flex-wrap: wrap; }
  .btn-confirm {
    padding: 8px 20px; font-size: 0.9rem; cursor: pointer;
    background: #1a7a3c; color: #fff; border: none; border-radius: 5px;
  }
  .btn-confirm:hover { background: #145e2e; }
  .btn-cancel {
    padding: 8px 16px; font-size: 0.9rem; cursor: pointer;
    background: #fff; color: #666; border: 1px solid #ccc; border-radius: 5px;
  }
  .btn-cancel:hover { background: #f0f0f0; }

  /* Divider */
  .divider { border: none; border-top: 1px solid #d0dce8; margin: 24px 0; }

  /* Add section */
  .add-toggle {
    background: none; border: none; cursor: pointer;
    font-size: 1rem; color: #2e7dd1; padding: 0; margin-bottom: 14px;
    display: flex; align-items: center; gap: 6px; font-weight: bold;
  }
  .add-toggle .arrow { transition: transform 0.2s; display: inline-block; }
  .add-toggle.open .arrow { transform: rotate(90deg); }
  .add-body { display: none; }
  .add-body.open { display: block; }
  .add-form {
    background: #fff; border: 1px solid #d0dce8; border-radius: 8px;
    padding: 18px 20px;
  }
  .form-row {
    display: flex; align-items: center; gap: 10px;
    margin-bottom: 12px; flex-wrap: wrap;
  }
  .form-row label { min-width: 80px; font-size: 0.9rem; color: #444; flex-shrink: 0; }
  .form-row input[type=text] {
    flex: 1; min-width: 160px; padding: 7px 10px; font-size: 0.95rem;
    border: 1px solid #c5d3e0; border-radius: 5px; outline: none;
  }
  .form-row input[type=text]:focus { border-color: #2e7dd1; }
  .form-row select {
    padding: 7px 8px; font-size: 0.9rem;
    border: 1px solid #c5d3e0; border-radius: 5px; outline: none; background: #fff;
  }
  .btn-add {
    padding: 9px 24px; font-size: 0.95rem; cursor: pointer;
    background: #2e7dd1; color: #fff; border: none; border-radius: 6px; margin-top: 4px;
  }
  .btn-add:hover { background: #1b5fa8; }

  /* User words list */
  .user-list { margin-top: 18px; }
  .user-list h3 { font-size: 0.95rem; color: #555; margin-bottom: 8px; }
  .user-word-item {
    display: flex; justify-content: space-between; align-items: center;
    padding: 6px 10px; background: #f7f0ff; border-radius: 5px; margin-bottom: 5px;
    font-size: 0.9rem;
  }
  .btn-del {
    padding: 3px 10px; font-size: 0.8rem; cursor: pointer;
    background: #fff; border: 1px solid #d0c0e0; border-radius: 4px; color: #800;
    flex-shrink: 0;
  }
  .btn-del:hover { background: #ffe8e8; }
  .server-status {
    font-size: 0.78rem; text-align: right; color: #aaa; margin-bottom: 6px;
  }
  .server-status.ok  { color: #1a7a3c; }
  .server-status.err { color: #c00; }
</style>
</head>
<body>
<div class="container">
  <h1>國中英文單字查詢</h1>
  <div class="server-status" id="serverStatus">正在連線至伺服器...</div>

  <div class="search-row">
    <input type="text" id="searchInput" placeholder="輸入英文單字（如: apple）"
           autocomplete="off" spellcheck="false">
    <button class="btn-primary" onclick="doSearch()">查詢</button>
  </div>

  <div id="resultArea"></div>

  <hr class="divider">

  <button class="add-toggle" id="addToggle" onclick="toggleAdd()">
    <span class="arrow">▶</span> 手動新增單字
  </button>
  <div class="add-body" id="addBody">
    <div class="add-form">
      <div class="form-row">
        <label>單字</label>
        <input type="text" id="addWord" placeholder="英文單字">
      </div>
      <div class="form-row">
        <label>KK音標</label>
        <input type="text" id="addKK" placeholder="如: ˈæpl">
      </div>
      <div class="form-row">
        <label>詞性 1</label>
        <select id="addPos1">
          <option value="n.">n. 名詞</option>
          <option value="v.">v. 動詞</option>
          <option value="adj.">adj. 形容詞</option>
          <option value="adv.">adv. 副詞</option>
          <option value="prep.">prep. 介系詞</option>
          <option value="conj.">conj. 連接詞</option>
          <option value="pron.">pron. 代名詞</option>
          <option value="interj.">interj. 感嘆詞</option>
          <option value="num.">num. 數詞</option>
        </select>
        <input type="text" id="addMeaning1" placeholder="中譯（必填）">
      </div>
      <div class="form-row">
        <label>詞性 2</label>
        <select id="addPos2">
          <option value="">（選填）</option>
          <option value="n.">n. 名詞</option>
          <option value="v.">v. 動詞</option>
          <option value="adj.">adj. 形容詞</option>
          <option value="adv.">adv. 副詞</option>
          <option value="prep.">prep. 介系詞</option>
          <option value="conj.">conj. 連接詞</option>
          <option value="pron.">pron. 代名詞</option>
          <option value="interj.">interj. 感嘆詞</option>
          <option value="num.">num. 數詞</option>
        </select>
        <input type="text" id="addMeaning2" placeholder="中譯（選填）">
      </div>
      <button class="btn-add" onclick="addWordManual()">新增</button>
    </div>
    <div class="user-list" id="userWordsList"></div>
  </div>
</div>

<script>
// ── Embedded vocabulary (2056 words) ────────────────────────────────────────
const DB = __DB_JSON__;

const POS_ZH = {
  'n.':'名詞','v.':'動詞','adj.':'形容詞','adv.':'副詞',
  'prep.':'介系詞','conj.':'連接詞','pron.':'代名詞',
  'interj.':'感嘆詞','num.':'數詞','art.':'冠詞'
};

const POS_SHORT = {
  'n.':'名','v.':'動','adj.':'形','adv.':'副',
  'prep.':'介','conj.':'連','pron.':'代','interj.':'感','num.':'數','art.':'冠',
  '名詞':'名','動詞':'動','形容詞':'形','副詞':'副',
  '介系詞':'介','連接詞':'連','代名詞':'代','感嘆詞':'感','數詞':'數','冠詞':'冠',
  '其他':'其'
};

// In-memory overlay: confirmed/manual words loaded from server
let userDB = {};
let serverOk = false;
let pendingAI = null;      // holds {word, data} while waiting for user confirmation
let currentCopyText = '';  // text ready to copy (avoids quoting issues in onclick)
const API = 'http://localhost:5001';  // absolute base — works from both file:// and http://

// ── Server communication ─────────────────────────────────────────────────────

async function initServer() {
  const st = document.getElementById('serverStatus');
  try {
    const resp = await fetch(API + '/api/user_words', {signal: AbortSignal.timeout(3000)});
    if (!resp.ok) throw new Error('bad status');
    userDB = await resp.json();
    serverOk = true;
    const n = Object.keys(userDB).length;
    st.textContent = '伺服器已連線' + (n ? `（已儲存 ${n} 筆補充單字）` : '');
    st.className = 'server-status ok';
    renderUserWordsList();
  } catch (e) {
    serverOk = false;
    st.textContent = '⚠ 伺服器未連線 — AI 補字功能暫停，請執行 python server.py';
    st.className = 'server-status err';
  }
}

// ── Search ────────────────────────────────────────────────────────────────────

function doSearch() {
  const raw = document.getElementById('searchInput').value.trim();
  if (!raw) return;
  const key = raw.toLowerCase();

  // 1. Check confirmed/manual user words
  if (userDB[key]) {
    const src = userDB[key].source === 'manual' ? '手動新增' : '已確認加入資料庫';
    const cls = userDB[key].source === 'manual' ? 'tag-user' : 'tag-confirmed';
    renderDbResult(raw, userDB[key], src, cls);
    return;
  }
  // 2. Check embedded DB
  if (DB[key]) {
    const src = DB[key].g ? '935 常用單字' : '補充（2000 單字表）';
    const cls = DB[key].g ? 'tag-935' : 'tag-supp';
    renderDbResult(raw, DB[key], src, cls);
    return;
  }
  // 3. Not found — try AI
  if (serverOk) {
    aiLookup(raw);
  } else {
    document.getElementById('resultArea').innerHTML =
      '<div class="result-card not-found">查無「' + escH(raw) +
      '」。請啟動伺服器（python server.py）後可由 AI 自動補充。</div>';
  }
}

// ── Render helpers ─────────────────────────────────────────────────────────────

function trimMean(mean) {
  var parts = String(mean).split('；');
  return parts.slice(0, 2).join('；');
}

function isDerivedForm(mean) {
  return /（[^）]*(?:現在分詞|過去式|過去分詞|比較級|最高級|第三人稱單數|名詞複數)[^）]*）/.test(String(mean));
}

function makePlainText(data) {
  var kk   = data.kk || '';
  var parts = ['[' + kk + ']'];

  if (data.entries && data.entries.length) {
    data.entries.filter(function(e) {
      return !isDerivedForm(e.mean);
    }).slice(0, 2).forEach(function(e) {
      var short = POS_SHORT[e.abbr] || POS_SHORT[e.zh] || e.abbr || '';
      parts.push('(' + short + ') ' + trimMean(e.mean));
    });
  } else {
    var t = data.t || '';
    var p = data.p || '';
    var short = POS_SHORT[p] || p;
    var m = t.match(/^\(([^)]+)\)\s*(.+)$/);
    var mean = trimMean(m ? m[2] : t);
    parts.push(short ? '(' + short + ') ' + mean : mean);
  }
  return parts.join(' ');
}

function renderDbResult(raw, data, sourceLabel, sourceClass) {
  pendingAI = null;
  var text = makePlainText(data);
  currentCopyText = text;
  document.getElementById('resultArea').innerHTML =
    '<div class="result-card found">' +
      '<div class="result-text">' + escH(text) + '</div>' +
      '<button class="copy-btn" id="copyBtnMain" onclick="copyText(currentCopyText)">複製</button>' +
      '<div class="source-tag">來源：<span class="' + sourceClass + '">' + sourceLabel + '</span></div>' +
    '</div>';
}

// ── AI lookup ─────────────────────────────────────────────────────────────────

async function aiLookup(word) {
  document.getElementById('resultArea').innerHTML =
    '<div class="result-card loading">AI 查詢「' + escH(word) + '」中...</div>';
  try {
    const resp = await fetch(API + '/api/lookup?word=' + encodeURIComponent(word));
    const json = await resp.json();
    if (json.found) {
      if (json.source === 'confirmed') {
        // Already confirmed before but not yet in userDB (e.g. page just loaded)
        userDB[word.toLowerCase()] = json.data;
        renderDbResult(word, json.data, '已確認加入資料庫', 'tag-confirmed');
      } else {
        renderAIPending(word, json.data);
      }
    } else {
      document.getElementById('resultArea').innerHTML =
        '<div class="result-card not-found">查無「' + escH(word) +
        '」（' + escH(json.error || '不明原因') + '）。請確認拼字，或手動新增。</div>';
    }
  } catch (e) {
    document.getElementById('resultArea').innerHTML =
      '<div class="result-card not-found">伺服器連線失敗：' + escH(e.message) + '</div>';
  }
}

function renderAIPending(word, data) {
  pendingAI = {word: word, data: data};
  var text = makePlainText(data);
  currentCopyText = text;
  document.getElementById('resultArea').innerHTML =
    '<div class="result-card ai-pending">' +
      '<div class="result-text">' + escH(text) + '</div>' +
      '<button class="copy-btn" id="copyBtnMain" onclick="copyText(currentCopyText)">複製</button>' +
      '<div class="source-tag">來源：<span class="tag-ai">AI 建議（待確認）</span></div>' +
      '<div class="confirm-row">' +
        '<button class="btn-confirm" onclick="confirmPending()">確認加入資料庫</button>' +
        '<button class="btn-cancel" onclick="document.getElementById(\'resultArea\').innerHTML=\'\'">取消</button>' +
      '</div>' +
    '</div>';
}

async function confirmPending() {
  if (!pendingAI) return;
  var word = pendingAI.word;
  var data = pendingAI.data;
  try {
    var resp = await fetch(API + '/api/confirm', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({word: word, data: data})
    });
    if (!resp.ok) throw new Error('HTTP ' + resp.status);
    userDB[word.toLowerCase()] = data;
    pendingAI = null;
    renderDbResult(word, data, '已確認加入資料庫 ✓', 'tag-confirmed');
    renderUserWordsList();
  } catch (e) {
    alert('儲存失敗：' + e.message);
  }
}

// ── Copy ──────────────────────────────────────────────────────────────────────

async function copyText(text) {
  var btn = document.getElementById('copyBtnMain');
  function feedback() {
    if (btn) { btn.textContent = '已複製！'; setTimeout(function(){ if(btn) btn.textContent = '複製'; }, 1500); }
  }
  try {
    // ClipboardItem 強制只寫 text/plain，不含任何 HTML 樣式
    var blob = new Blob([text], {type: 'text/plain'});
    await navigator.clipboard.write([new ClipboardItem({'text/plain': blob})]);
    feedback();
  } catch (e) {
    // 舊瀏覽器 fallback：隱藏 textarea，同樣只有純文字
    var ta = document.createElement('textarea');
    ta.value = text;
    ta.style.cssText = 'position:fixed;top:-9999px;left:-9999px;opacity:0';
    document.body.appendChild(ta);
    ta.focus(); ta.select();
    try { document.execCommand('copy'); } catch(e2) {}
    document.body.removeChild(ta);
    feedback();
  }
}

function escH(str) {
  return String(str).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}

// ── Manual add ────────────────────────────────────────────────────────────────

function toggleAdd() {
  document.getElementById('addToggle').classList.toggle('open');
  document.getElementById('addBody').classList.toggle('open');
}

async function addWordManual() {
  var word  = document.getElementById('addWord').value.trim();
  var kk    = document.getElementById('addKK').value.trim();
  var abbr1 = document.getElementById('addPos1').value;
  var mean1 = document.getElementById('addMeaning1').value.trim();
  var abbr2 = document.getElementById('addPos2').value;
  var mean2 = document.getElementById('addMeaning2').value.trim();

  if (!word || !kk || !mean1) { alert('請填入單字、KK音標與詞性1中譯'); return; }

  var entries = [{abbr: abbr1, zh: POS_ZH[abbr1] || '', mean: mean1}];
  if (abbr2 && mean2) entries.push({abbr: abbr2, zh: POS_ZH[abbr2] || '', mean: mean2});

  var data = {w: word, kk: kk, entries: entries, source: 'manual'};

  if (serverOk) {
    try {
      var resp = await fetch(API + '/api/confirm', {
        method: 'POST', headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({word: word, data: data})
      });
      if (!resp.ok) throw new Error('HTTP ' + resp.status);
    } catch (e) { alert('儲存失敗：' + e.message); return; }
  }

  userDB[word.toLowerCase()] = data;
  document.getElementById('addWord').value = '';
  document.getElementById('addKK').value = '';
  document.getElementById('addMeaning1').value = '';
  document.getElementById('addPos2').value = '';
  document.getElementById('addMeaning2').value = '';
  alert('已新增「' + word + '」' + (serverOk ? '並儲存至伺服器' : '（僅暫存，請啟動伺服器以永久儲存）'));
  renderUserWordsList();
}

// ── User words list ───────────────────────────────────────────────────────────

function renderUserWordsList() {
  var list = document.getElementById('userWordsList');
  var keys = Object.keys(userDB);
  if (!keys.length) { list.innerHTML = ''; return; }

  var rows = keys.map(function(k) {
    var d = userDB[k];
    var summary = '';
    if (d.entries) {
      summary = d.entries.map(function(e){ return '(' + e.abbr + ') ' + e.mean; }).join(' / ');
    } else {
      summary = d.t || '';
    }
    return '<div class="user-word-item">' +
      '<span>' + escH(d.w) + '&nbsp;&nbsp;[' + escH(d.kk) + ']&nbsp;&nbsp;' + escH(summary) + '</span>' +
      '<button class="btn-del" onclick="deleteWord(\'' + escH(k) + '\')">刪除</button>' +
    '</div>';
  });
  list.innerHTML = '<h3>補充單字（共 ' + keys.length + ' 筆）</h3>' + rows.join('');
}

async function deleteWord(key) {
  if (!confirm('確定刪除「' + key + '」？')) return;
  if (serverOk) {
    try {
      await fetch(API + '/api/delete', {
        method: 'POST', headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({word: key})
      });
    } catch (e) { alert('刪除失敗：' + e.message); return; }
  }
  delete userDB[key];
  renderUserWordsList();
}

// ── Init ──────────────────────────────────────────────────────────────────────

document.getElementById('searchInput').addEventListener('keypress', function(e) {
  if (e.key === 'Enter') doSearch();
});

initServer();
</script>
</body>
</html>
"""

html = HTML.replace('__DB_JSON__', db_json)

with open('index.html', 'w', encoding='utf-8') as f:
    f.write(html)

print(f'index.html generated  ({len(html):,} bytes, {len(db)} words)')
